from flask import Blueprint, render_template, session, redirect, url_for, request, flash, send_file, current_app
from utils.decorators import role_required
from database import db
from models.user import User
from models.course import Course, CourseStudent
from models.schedule import Schedule
from models.attendance_session import AttendanceSession
from models.attendance_record import AttendanceRecord
from services import statistics_service, export_service
from services.attendance_service import (
    get_active_session, start_session as attendance_start_session, get_session_by_id,
    is_code_expired, refresh_code, get_session_records,
    get_suspicious_records, get_enrolled_count, end_session as attendance_end_session,
    resolve_suspicious as attendance_resolve_suspicious,
)
from utils.qr_generator import generate_qr_base64

teacher_bp = Blueprint('teacher', __name__)


@teacher_bp.route('/dashboard')
@role_required(1)
def dashboard():
    user_id = session['user']['id']
    my_courses = db.query(Course).filter(Course.teacher_id == user_id, Course.is_active == 1).all()
    
    active_sessions = {}
    for course in my_courses:
        active = get_active_session(course.id)
        if active:
            active_sessions[course.id] = active.id

    # Derslerin program bilgilerini al
    course_schedules = {}
    for course in my_courses:
        schedules = db.query(Schedule).filter_by(course_id=course.id).all()
        course_schedules[course.id] = schedules

    # Ders programındaki en erken ve en geç saatleri bul
    all_times = []
    for schedules in course_schedules.values():
        for schedule in schedules:
            all_times.append(schedule.start_time)
            all_times.append(schedule.end_time)
    
    if all_times:
        # Saatleri parse et ve min/max bul (güvenli parsing)
        start_times = []
        for t in all_times:
            try:
                t_str = str(t) if t else "00:00"
                if ':' not in t_str:
                    t_str = f"{int(float(t_str)):02d}:00"
                start_times.append(int(t_str.split(':')[0]))
            except:
                start_times.append(0)  # Hata olursa varsayılan
        
        if start_times:
            min_hour = min(start_times)
            max_hour = max(start_times)
        else:
            min_hour = 8
            max_hour = 18
        
        # 1 saat önce başla, 1 saat sonra bitir
        start_hour = max(7, min_hour - 1)
        end_hour = min(22, max_hour + 1)
    else:
        # Varsayılan saatler
        start_hour = 8
        end_hour = 18
    
    # Saat dilimlerini oluştur (30 dakika aralıklarla)
    time_slots = []
    for hour in range(start_hour, end_hour):
        time_slots.append(f"{hour:02d}:00")
        time_slots.append(f"{hour:02d}:30")

    # Her ders için renk ataması yap
    course_colors = {}
    color_palette = [
        '#3498db', '#27ae60', '#e74c3c', '#f39c12', '#9b59b6',
        '#1abc9c', '#34495e', '#e67e22', '#16a085', '#2c3e50'
    ]
    
    for i, course in enumerate(my_courses):
        course_colors[course.id] = color_palette[i % len(color_palette)]

    return render_template('teacher/dashboard.html', 
                         courses=my_courses, 
                         active_sessions=active_sessions, 
                         course_schedules=course_schedules,
                         time_slots=time_slots,
                         course_colors=course_colors)


@teacher_bp.route('/course/<int:course_id>/details')
@role_required(1)
def course_details(course_id):
    user_id = session['user']['id']
    course = db.query(Course).filter_by(id=course_id, teacher_id=user_id).first()
    
    if not course:
        flash('Ders bulunamadı veya yetkiniz yok.', 'error')
        return redirect(url_for('teacher.dashboard'))
    
    # Dersin öğrencileri
    course_students = db.query(CourseStudent).filter_by(course_id=course_id).all()
    student_ids = [cs.student_id for cs in course_students]
    students = db.query(User).filter(User.id.in_(student_ids), User.is_active == 1).all()
    
    # Aktif yoklama oturumu
    active_session = get_active_session(course_id)
    
    # Dersin programı
    schedules = db.query(Schedule).filter_by(course_id=course_id).all()
    
    return render_template('teacher/course_details.html', 
                         course=course, 
                         students=students, 
                         active_session=active_session,
                         schedules=schedules)


@teacher_bp.route('/course/<int:course_id>/student/<int:student_id>/attendance')
@role_required(1)
def student_attendance(course_id, student_id):
    try:
        user_id = session['user']['id']
        course = db.query(Course).filter_by(id=course_id, teacher_id=user_id).first()
        
        if not course:
            flash('Ders bulunamadı veya yetkiniz yok.', 'error')
            return redirect(url_for('teacher.dashboard'))
        
        student = db.query(User).filter_by(id=student_id).first()
        if not student:
            flash('Öğrenci bulunamadı.', 'error')
            return redirect(url_for('teacher.course_details', course_id=course_id))
        
        # Öğrencinin bu dersteki yoklama geçmişini al
        attendance_records = db.query(AttendanceRecord).join(AttendanceSession).filter(
            AttendanceSession.course_id == course_id,
            AttendanceRecord.student_id == student_id
        ).order_by(AttendanceSession.started_at.desc()).all()
        
        return render_template('teacher/student_attendance.html',
                             course=course,
                             student=student,
                             attendance_records=attendance_records)
    
    except Exception as e:
        flash(f'Öğrenci yoklama geçmişi yüklenirken hata: {str(e)}', 'error')
        return redirect(url_for('teacher.course_details', course_id=course_id))


@teacher_bp.route('/course/<int:course_id>/schedule', methods=['GET', 'POST'])
@role_required(1)
def course_schedule(course_id):
    try:
        user_id = session['user']['id']
        course = db.query(Course).filter_by(id=course_id, teacher_id=user_id).first()
        if not course:
            flash('Ders bulunamadı.', 'error')
            return redirect(url_for('teacher.dashboard'))

        if request.method == 'POST':
            day_of_week = request.form.get('day_of_week', type=int)
            start_time = request.form.get('start_time', '').strip()
            end_time = request.form.get('end_time', '').strip()
            room = request.form.get('room', '').strip()

            if day_of_week is not None and start_time and end_time:
                schedule = Schedule(
                    course_id=course_id,
                    day_of_week=day_of_week,
                    start_time=start_time,
                    end_time=end_time,
                    room=room or None,
                )
                db.add(schedule)
                db.commit()
                flash('Ders saati eklendi.', 'success')
            else:
                flash('Tüm alanlar doldurulmalıdır.', 'error')

        schedules = db.query(Schedule).filter_by(course_id=course_id).order_by(Schedule.day_of_week).all()
        active_session = get_active_session(course_id)
        
        # Öğretmenin tüm derslerini ve programlarını al
        teacher_courses = db.query(Course).filter(Course.teacher_id == user_id, Course.is_active == 1).all()
        all_schedules = []
        for course in teacher_courses:
            course_schedules = db.query(Schedule).filter_by(course_id=course.id).all()
            for s in course_schedules:
                all_schedules.append({
                    'schedule': s,
                    'course': course
                })
        
        days = ['Pazartesi', 'Salı', 'Çarşamba', 'Perşembe', 'Cuma']
        time_slots = ['08:00', '08:30', '09:00', '09:30', '10:00', '10:30', '11:00', '11:30',
                      '12:00', '12:30', '13:00', '13:30', '14:00', '14:30', '15:00', '15:30',
                      '16:00', '16:30', '17:00', '17:30', '18:00', '18:30']

        # {gün: {saat: ders_kodu}} şeklinde sözlük
        schedule_data = {day: {slot: '' for slot in time_slots} for day in days}

        for item in all_schedules:
            s = item['schedule']
            course = item['course']
            if s.day_of_week is None or s.day_of_week >= len(days):
                continue
            day_name = days[s.day_of_week]
            try:
                def _to_str(val):
                    v = str(val) if val else '00:00'
                    if ':' not in v:
                        v = f"{int(float(v)):02d}:00"
                    return v
                start_str = _to_str(s.start_time)
                end_str = _to_str(s.end_time)
                start_min = int(start_str[:2]) * 60 + int(start_str[3:5])
                end_min = int(end_str[:2]) * 60 + int(end_str[3:5])
                for slot in time_slots:
                    slot_min = int(slot[:2]) * 60 + int(slot[3:5])
                    if start_min <= slot_min < end_min:
                        cur = schedule_data[day_name][slot]
                        code = course.code or 'KOD'
                        schedule_data[day_name][slot] = f"{cur}, {code}" if cur else code
            except Exception:
                continue
        
        return render_template('teacher/schedule.html', 
                             course=course, 
                             schedules=schedules, 
                             active_session=active_session,
                             schedule_data=schedule_data,
                             all_courses=teacher_courses)
        
    except Exception as e:
        flash(f'Ders programı yüklenirken hata: {str(e)}', 'error')
        return redirect(url_for('teacher.dashboard'))


@teacher_bp.route('/attendance/start')
@role_required(1)
def attendance_start():
    """Öğretmene atanan dersler arasından yoklama başlatma sayfası"""
    user_id = session['user']['id']
    course_id = request.args.get('course_id', type=int)
    
    # Eğer course_id varsa, direkt QR ayarları sayfasına git
    if course_id:
        course = db.query(Course).filter_by(id=course_id, teacher_id=user_id, is_active=1).first()
        if not course:
            flash('Ders bulunamadı veya yetkiniz yok.', 'error')
            return redirect(url_for('teacher.dashboard'))
        
        # Aktif oturum kontrolü
        active_session = get_active_session(course_id)
        if active_session:
            return redirect(url_for('teacher.active_session', session_id=active_session.id))
        
        # Tek ders için veri hazırla
        course_data = {
            'course': {
                'id': course.id,
                'name': course.name,
                'code': course.code,
                'department': course.department,
                'class_name': course.class_name
            },
            'schedules': [],
            'active_session': None
        }
        
        return render_template('teacher/attendance_start.html', 
                           courses_data=[course_data], 
                           preselected_course=True)
    
    # course_id yoksa, tüm dersleri göster (normal akış)
    # Öğretmenin tüm derslerini al (onay filtresi geçici olarak kaldırıldı)
    teacher_courses = db.query(Course).filter(
        Course.teacher_id == user_id, 
        Course.is_active == 1
    ).all()
    
    # Her dersin programını ve aktif oturumunu al
    courses_data = []
    for course in teacher_courses:
        schedules = db.query(Schedule).filter_by(course_id=course.id).all()
        active_session = get_active_session(course.id)
        
        courses_data.append({
            'course': {
                'id': course.id,
                'name': course.name,
                'code': course.code,
                'department': course.department,
                'class_name': course.class_name
            },
            'schedules': [{
                'id': s.id,
                'day_of_week': s.day_of_week,
                'start_time': s.start_time,
                'end_time': s.end_time,
                'room': s.room
            } for s in schedules],
            'active_session': {
                'id': active_session.id,
                'status': active_session.status
            } if active_session else None
        })
    
    return render_template('teacher/attendance_start.html', courses_data=courses_data, preselected_course=False)


@teacher_bp.route('/course/<int:course_id>/start-session', methods=['POST'])
@role_required(1)
def start_session(course_id):
    user_id = session['user']['id']
    refresh_seconds = request.form.get('refresh_seconds', type=int) or 3
    refresh_min = current_app.config.get('QR_REFRESH_MIN', 1)
    refresh_max = current_app.config.get('QR_REFRESH_MAX', 60)
    refresh_seconds = max(refresh_min, min(refresh_seconds, refresh_max))
    allowed_ip = request.form.get('allowed_ip_prefix', '').strip() or current_app.config.get('ALLOWED_IP_PREFIX')
    latitude = request.form.get('latitude', type=float)
    longitude = request.form.get('longitude', type=float)
    radius_m = request.form.get('radius_m', type=int)
    
    # Çoklu saat seçimi desteği
    schedule_ids = request.form.getlist('schedule_ids')
    schedule_id = None
    if schedule_ids:
        schedule_id = int(schedule_ids[0])  # İlk seçili saat
    
    # Formdaki course_id'yi kontrol et
    form_course_id = request.form.get('course_id', type=int)
    if form_course_id and form_course_id != course_id:
        course_id = form_course_id

    att_session, error = attendance_start_session(
        course_id=course_id,
        teacher_id=user_id,
        schedule_id=schedule_id,
        refresh_seconds=refresh_seconds,
        allowed_ip_prefix=allowed_ip,
        latitude=latitude,
        longitude=longitude,
        radius_m=radius_m
    )

    if error:
        flash(error, 'error')
        return redirect(url_for('teacher.attendance_start'))

    flash('Yoklama oturumu başarıyla başlatıldı.', 'success')
    return redirect(url_for('teacher.active_session', session_id=att_session.id))


@teacher_bp.route('/session/<session_id>')
@role_required(1)
def active_session(session_id):
    user_id = session['user']['id']
    att_session = get_session_by_id(session_id)
    if not att_session or att_session.teacher_id != user_id:
        flash('Oturum bulunamadı.', 'error')
        return redirect(url_for('teacher.dashboard'))

    if att_session.status == 'active' and is_code_expired(att_session):
        att_session = refresh_code(att_session.id)

    records = get_session_records(session_id)
    suspicious_records = get_suspicious_records(session_id)
    enrolled_count = get_enrolled_count(att_session.course_id)

    qr_data = att_session.current_code
    qr_base64 = generate_qr_base64(qr_data) if att_session.status == 'active' else None

    return render_template('teacher/active_session.html',
                           att_session=att_session,
                           records=records,
                           suspicious_records=suspicious_records,
                           enrolled_count=enrolled_count,
                           qr_base64=qr_base64)


@teacher_bp.route('/session/<session_id>/update-refresh', methods=['POST'])
@role_required(1)
def update_refresh_rate(session_id):
    user_id = session['user']['id']
    att_session = get_session_by_id(session_id)
    
    if not att_session or att_session.teacher_id != user_id:
        flash('Oturum bulunamadı veya yetkiniz yok.', 'error')
        return redirect(url_for('teacher.dashboard'))
    
    if att_session.status != 'active':
        flash('Sadece aktif oturumların yenileme süresi güncellenebilir.', 'error')
        return redirect(url_for('teacher.active_session', session_id=session_id))
    
    refresh_seconds = request.form.get('refresh_seconds', type=int)
    if not refresh_seconds or refresh_seconds < 1 or refresh_seconds > 60:
        flash('Geçersiz yenileme süresi. 1-60 saniye arası olmalı.', 'error')
        return redirect(url_for('teacher.active_session', session_id=session_id))
    
    # Oturumun yenileme süresini güncelle
    att_session.code_refresh_seconds = refresh_seconds
    db.commit()
    
    flash('QR yenileme süresi güncellendi.', 'success')
    return redirect(url_for('teacher.active_session', session_id=session_id))


@teacher_bp.route('/session/<session_id>/end', methods=['POST'])
@role_required(1)
def end_session(session_id):
    user_id = session['user']['id']
    att_session, error = attendance_end_session(session_id, user_id)
    if error:
        flash(error, 'error')
        return redirect(url_for('teacher.dashboard'))

    flash('Yoklama oturumu sonlandırıldı.', 'success')
    return redirect(url_for('teacher.active_session', session_id=session_id))


@teacher_bp.route('/records/<int:record_id>/resolve', methods=['POST'])
@role_required(1)
def resolve_suspicious(record_id):
    user_id = session['user']['id']
    decision = request.form.get('decision', '').strip()
    note = request.form.get('note', '').strip() or None

    record, error = attendance_resolve_suspicious(record_id, user_id, decision, note)
    if error:
        flash(error, 'error')
        return redirect(url_for('teacher.dashboard'))

    if record.status == 'approved':
        flash('Şüpheli yoklama onaylandı.', 'success')
    else:
        flash('Şüpheli yoklama reddedildi.', 'warning')
    return redirect(url_for('teacher.active_session', session_id=record.session_id))


@teacher_bp.route('/schedule_all')
@role_required(1)
def schedule_all():
    user_id = session['user']['id']
    teacher = db.query(User).filter(User.id == user_id).first()
    schedules = db.query(Schedule).join(Course).filter(Course.teacher_id == user_id).all()
    return render_template('teacher/schedule_all.html', schedules=schedules, teacher=teacher)


@teacher_bp.route('/course_approvals')
@role_required(1)
def course_approvals():
    """Öğretmenin onay bekleyen dersleri"""
    user_id = session['user']['id']
    # Bu öğretmene atanmış ama onay bekleyen dersler
    pending_courses = db.query(Course).filter(
        Course.teacher_id == user_id, 
        Course.teacher_approval == 0,
        Course.status == 0
    ).all()
    
    # Onaylanan ve reddedilen dersler
    approved_courses = db.query(Course).filter(
        Course.teacher_id == user_id, 
        Course.teacher_approval == 1
    ).all()
    
    rejected_courses = db.query(Course).filter(
        Course.teacher_id == user_id, 
        Course.teacher_approval == 2
    ).all()
    
    return render_template('teacher/course_approvals.html', 
                         pending_courses=pending_courses,
                         approved_courses=approved_courses,
                         rejected_courses=rejected_courses)


@teacher_bp.route('/approve_course/<int:course_id>', methods=['POST'])
@role_required(1)
def approve_course(course_id):
    """Dersi onayla"""
    user_id = session['user']['id']
    course = db.query(Course).filter(
        Course.id == course_id, 
        Course.teacher_id == user_id
    ).first()
    
    if not course:
        flash('Ders bulunamadı.', 'error')
        return redirect(url_for('teacher.course_approvals'))
    
    action = request.form.get('action')
    if action == 'approve':
        course.teacher_approval = 1
        course.status = 1  # Aktif yap
        flash('Ders başarıyla onaylandı.', 'success')
    elif action == 'reject':
        course.teacher_approval = 2
        course.status = 2  # Pasif yap
        flash('Ders reddedildi.', 'warning')
    
    db.commit()
    return redirect(url_for('teacher.course_approvals'))


@teacher_bp.route('/student_approvals')
@role_required(1)
def student_approvals():
    """Öğretmenin derse eklediği ama admin onayı bekleyen öğrenciler"""
    user_id = session['user']['id']
    
    # Bu öğretmenin derslerinde admin onayı bekleyen öğrenciler
    pending_students = db.query(CourseStudent).join(Course).filter(
        Course.teacher_id == user_id,
        CourseStudent.admin_approval == 0
    ).all()
    
    # Admin onayı almış öğrenciler
    approved_students = db.query(CourseStudent).join(Course).filter(
        Course.teacher_id == user_id,
        CourseStudent.admin_approval == 1
    ).all()
    
    return render_template('teacher/student_approvals.html',
                         pending_students=pending_students,
                         approved_students=approved_students)


@teacher_bp.route('/statistics')
@role_required(1)
def statistics():
    user_id = session['user']['id']
    course_id = request.args.get('course_id', type=int)
    
    # Öğretmenin derslerini al
    teacher_courses = db.query(Course).filter(Course.teacher_id == user_id, Course.is_active == 1).all()
    
    if course_id:
        # Seçili dersin istatistikleri
        course = db.query(Course).filter_by(id=course_id, teacher_id=user_id).first()
        if not course:
            flash('Ders bulunamadı veya yetkiniz yok.', 'error')
            return redirect(url_for('teacher.statistics'))
        
        # Seçili ders için istatistikleri oluştur
        all_stats = statistics_service.get_teacher_statistics(user_id)
        # Sadece seçili dersin istatistiklerini filtrele
        filtered_course_stats = [cs for cs in all_stats.get('course_stats', []) if cs['course'].id == course_id]
        
        # Öğrenci devamsızlık verilerini al
        course_students = db.query(CourseStudent).filter_by(course_id=course_id).all()
        student_ids = [cs.student_id for cs in course_students]
        students = db.query(User).filter(User.id.in_(student_ids), User.is_active == 1).all()
        
        student_absences = []
        for student in students:
            # Öğrencinin bu dersteki yoklama kayıtlarını al
            try:
                attendance_records = db.query(AttendanceRecord).join(AttendanceSession).filter(
                    AttendanceSession.course_id == course_id,
                    AttendanceRecord.student_id == student.id
                ).all()
                
                total_sessions = db.query(AttendanceSession).filter_by(course_id=course_id).count()
                present_statuses = statistics_service.PRESENT_STATUSES
                present_count = len([ar for ar in attendance_records if ar.status in present_statuses])
                suspicious_count = len([ar for ar in attendance_records if ar.status == 'suspicious'])
                absence_count = max(total_sessions - present_count - suspicious_count, 0)
            except Exception as e:
                attendance_records = []
                absence_count = 0
                total_sessions = 0
            
            student_absences.append({
                'student': student,
                'absence_count': absence_count,
                'total_sessions': total_sessions,
                'attendance_rate': round((total_sessions - absence_count) / total_sessions * 100, 1) if total_sessions > 0 else 0
            })
        
        stats = {
            'course_stats': filtered_course_stats,
            'selected_course': course,
            'teacher_courses': teacher_courses,
            'selected_course_id': course_id,
            'student_absences': student_absences
        }
    else:
        # Tüm derslerin genel istatistikleri
        stats = statistics_service.get_teacher_statistics(user_id)
        stats['selected_course'] = None
        stats['teacher_courses'] = teacher_courses
        stats['selected_course_id'] = None
    
        
    return render_template('teacher/statistics.html', **stats)


@teacher_bp.route('/export/course/<int:course_id>')
@role_required(1)
def export_course(course_id):
    user_id = session['user']['id']
    course = db.query(Course).filter_by(id=course_id, teacher_id=user_id).first()
    if not course:
        flash('Ders bulunamadı veya yetkiniz yok.', 'error')
        return redirect(url_for('teacher.statistics'))
    result, filename_or_error = export_service.export_course_attendance(course_id)
    if result is None:
        flash(filename_or_error, 'error')
        return redirect(url_for('teacher.statistics'))
    return send_file(result, download_name=filename_or_error, as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@teacher_bp.route('/export/schedule')
@role_required(1)
def export_schedule():
    try:
        user_id = session['user']['id']
        
        # Öğretmenin derslerini al
        my_courses = db.query(Course).filter(Course.teacher_id == user_id, Course.is_active == 1).all()
        
        # Derslerin program bilgilerini al
        course_schedules = {}
        for course in my_courses:
            schedules = db.query(Schedule).filter_by(course_id=course.id).all()
            course_schedules[course.id] = schedules
        
        # Saat dilimlerini oluştur
        time_slots = []
        start_hour = 9
        end_hour = 18
        for hour in range(start_hour, end_hour):
            time_slots.append(f"{hour:02d}:00")
            time_slots.append(f"{hour:02d}:30")
        
        # Excel dosyası oluştur (openpyxl ile)
        import io
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from datetime import datetime
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Ders Programı"
            
            # Başlıkları yaz
            days = ['Saat', 'Pazartesi', 'Salı', 'Çarşamba', 'Perşembe', 'Cuma', 'Cumartesi', 'Pazar']
            
            # Başlık formatı
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="3498db", end_color="3498db", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                           top=Side(style='thin'), bottom=Side(style='thin'))
            
            # Başlıkları ekle
            for col, day in enumerate(days, 1):
                cell = ws.cell(row=1, column=col, value=day)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
                ws.column_dimensions[chr(64 + col)].width = 15
            
            # Saat formatı
            time_font = Font(bold=True)
            time_fill = PatternFill(start_color="ecf0f1", end_color="ecf0f1", fill_type="solid")
            
            # Tabloyu oluştur (DataFrame mantığı ile)
            days_map = {1: 'Pazartesi', 2: 'Salı', 3: 'Çarşamba', 4: 'Perşembe', 5: 'Cuma'}
            schedule_dict = {}
            
            # Tüm saat dilimleri için boş satırlar oluştur
            for time_slot in time_slots:
                schedule_dict[time_slot] = {day: '' for day in days[1:]}
            
            # Dersleri yerleştir
            for course_id, schedules in course_schedules.items():
                course = next((c for c in my_courses if c.id == course_id), None)
                if not course:
                    continue
                    
                for schedule in schedules:
                    day_name = days_map.get(schedule.day_of_week)
                    if day_name:
                        # Başlangıç ve bitiş saatlerini dakikaya çevir
                        start_minutes = int(schedule.start_time[:2]) * 60 + int(schedule.start_time[3:5])
                        end_minutes = int(schedule.end_time[:2]) * 60 + int(schedule.end_time[3:5])
                        
                        # Bu aralıktaki tüm time slot'ları doldur
                        for time_slot in time_slots:
                            slot_minutes = int(time_slot[:2]) * 60 + int(time_slot[3:5])
                            if slot_minutes >= start_minutes and slot_minutes < end_minutes:
                                if time_slot in schedule_dict and day_name in schedule_dict[time_slot]:
                                    schedule_dict[time_slot][day_name] = f"{course.name} ({schedule.room or '-'})"
            
            # Excel satırlarını yaz
            for row_idx, time_slot in enumerate(time_slots, 2):
                # Saat sütunu
                time_cell = ws.cell(row=row_idx, column=1, value=time_slot)
                time_cell.font = time_font
                time_cell.fill = time_fill
                time_cell.alignment = Alignment(horizontal="center", vertical="center")
                time_cell.border = border
                
                # Gün sütunları
                for col_idx, day in enumerate(days[1:], 2):
                    course_text = schedule_dict.get(time_slot, {}).get(day, '')
                    cell = ws.cell(row=row_idx, column=col_idx, value=course_text)
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    cell.border = border
                    
                    if course_text:
                        cell.fill = PatternFill(start_color="3498db", end_color="3498db", fill_type="solid")
                        cell.font = Font(color="FFFFFF", bold=True)
            
            # Dosyayı kaydet
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            filename = f"ders_programi_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            return send_file(output, download_name=filename, as_attachment=True,
                             mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        
        except ImportError:
            # openpyxl yüklü değilse CSV formatında döndür
            import csv
            from datetime import datetime
            
            output = io.StringIO()
            writer = csv.writer(output)
            
            # Başlıkları yaz
            days = ['Saat', 'Pazartesi', 'Salı', 'Çarşamba', 'Perşembe', 'Cuma', 'Cumartesi', 'Pazar']
            writer.writerow(days)
            
            # Tabloyu oluştur (DataFrame mantığı ile)
            days_map = {1: 'Pazartesi', 2: 'Salı', 3: 'Çarşamba', 4: 'Perşembe', 5: 'Cuma'}
            schedule_dict = {}
            
            # Tüm saat dilimleri için boş satırlar oluştur
            for time_slot in time_slots:
                schedule_dict[time_slot] = {day: '' for day in days[1:]}
            
            # Dersleri yerleştir
            for course_id, schedules in course_schedules.items():
                course = next((c for c in my_courses if c.id == course_id), None)
                if not course:
                    continue
                    
                for schedule in schedules:
                    day_name = days_map.get(schedule.day_of_week)
                    if day_name:
                        # Başlangıç ve bitiş saatlerini dakikaya çevir
                        start_minutes = int(schedule.start_time[:2]) * 60 + int(schedule.start_time[3:5])
                        end_minutes = int(schedule.end_time[:2]) * 60 + int(schedule.end_time[3:5])
                        
                        # Bu aralıktaki tüm time slot'ları doldur
                        for time_slot in time_slots:
                            slot_minutes = int(time_slot[:2]) * 60 + int(time_slot[3:5])
                            if slot_minutes >= start_minutes and slot_minutes < end_minutes:
                                if time_slot in schedule_dict and day_name in schedule_dict[time_slot]:
                                    schedule_dict[time_slot][day_name] = f"{course.name} ({schedule.room or '-'})"
            
            # CSV satırlarını yaz
            for time_slot, row_data in schedule_dict.items():
                csv_row = [time_slot] + [row_data.get(day, '') for day in days[1:]]
                writer.writerow(csv_row)
            
            # Dosyayı oluştur
            csv_content = output.getvalue()
            output_bytes = io.BytesIO(csv_content.encode('utf-8'))
            
            filename = f"ders_programi_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
            
            return send_file(output_bytes, download_name=filename, as_attachment=True,
                             mimetype='text/csv')
    
    except Exception as e:
        flash(f'Excel oluşturulurken hata: {str(e)}', 'error')
        return redirect(url_for('teacher.dashboard'))


@teacher_bp.route('/schedule/<int:schedule_id>/edit', methods=['GET', 'POST'])
@role_required(1)
def edit_schedule(schedule_id):
    user_id = session['user']['id']
    
    # Schedule'ı bul ve öğretmenin kendi dersinin olduğunu kontrol et
    schedule = db.query(Schedule).join(Course).filter(
        Schedule.id == schedule_id,
        Course.teacher_id == user_id
    ).first()
    
    if not schedule:
        flash('Program bulunamadı veya yetkiniz yok.', 'error')
        return redirect(url_for('teacher.dashboard'))
    
    if request.method == 'POST':
        # Form verilerini güncelle
        schedule.day_of_week = request.form.get('day_of_week', type=int)
        schedule.start_time = request.form.get('start_time')
        schedule.end_time = request.form.get('end_time')
        schedule.room = request.form.get('room')
        
        db.commit()
        flash('Ders programı güncellendi.', 'success')
        return redirect(url_for('teacher.course_details', course_id=schedule.course_id))
    
    return render_template('teacher/edit_schedule.html', schedule=schedule)
