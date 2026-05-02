import os
import sys
from datetime import datetime, timedelta
from io import BytesIO

from openpyxl import load_workbook

ROOT_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
sys.path.insert(0, ROOT_DIR)

from app import create_app
from database import db
from models.attendance_record import AttendanceRecord
from models.attendance_session import AttendanceSession
from models.course import Course, CourseStudent
from models.device_pairing import DevicePairing
from models.user import User
from services import attendance_service
from utils.hashing import hash_password


def create_fixture():
    admin = User(username='admin', email='admin@test.local', hashed_password=hash_password('pass'), role=0)
    teacher = User(username='teacher', email='teacher@test.local', hashed_password=hash_password('pass'), role=1)
    other_teacher = User(username='other', email='other@test.local', hashed_password=hash_password('pass'), role=1)
    student = User(
        username='student',
        email='student@test.local',
        hashed_password=hash_password('pass'),
        role=2,
        student_number='9001',
        department='Bilgisayar Mühendisliği',
    )
    db.add_all([admin, teacher, other_teacher, student])
    db.commit()

    course = Course(name='Entegrasyon Dersi', code='INT101', teacher_id=teacher.id, department='Bilgisayar Mühendisliği')
    db.add(course)
    db.commit()
    db.add(CourseStudent(course_id=course.id, student_id=student.id))
    db.commit()

    att_session = AttendanceSession(
        course_id=course.id,
        teacher_id=teacher.id,
        current_code='ABC123',
        code_expires_at=(datetime.utcnow() + timedelta(minutes=5)).isoformat(),
        code_refresh_seconds=10,
        allowed_ip_prefix='10.0.',
    )
    db.add(att_session)
    db.commit()
    return admin, teacher, other_teacher, student, course, att_session


def login(client, username):
    response = client.post('/login', data={'username': username, 'password': 'pass'})
    assert response.status_code in (302, 303), (username, response.status_code)


def pair_student_device(student):
    db.add(DevicePairing(student.id, 'DV:TESTDEVICE01', student.student_number))
    db.commit()


def run():
    pairing_optional_app = create_app('testing')
    pairing_optional_app.config['REQUIRE_DEVICE_PAIRING'] = False
    with pairing_optional_app.app_context():
        _, _, _, optional_student, _, optional_session = create_fixture()
        optional_client = pairing_optional_app.test_client()
        login(optional_client, 'student')
        optional_check_in = optional_client.post(
            f'/student/session/{optional_session.id}/check-in',
            data={'code': 'ABC123', 'override': '1', 'override_reason': 'Demo cihaz eşleşmesi kapalı'},
            environ_base={'REMOTE_ADDR': '8.8.8.8'},
        )
        assert optional_check_in.status_code in (302, 303)
        optional_record = db.query(AttendanceRecord).filter_by(
            session_id=optional_session.id,
            student_id=optional_student.id,
        ).first()
        assert optional_record is not None

    app = create_app('testing')
    with app.app_context():
        admin, teacher, other_teacher, student, course, att_session = create_fixture()
        client = app.test_client()

        assert client.get('/login').status_code == 200
        assert client.get('/sw.js').status_code == 200
        assert client.get('/static/js/offline.js').status_code == 200
        assert 'offline.js' in client.get('/login').get_data(as_text=True)

        login(client, 'student')
        dashboard = client.get('/student/dashboard')
        assert dashboard.status_code == 200
        dashboard_html = dashboard.get_data(as_text=True)
        assert 'data-offline-attendance' in dashboard_html
        assert 'BarcodeDetector' in dashboard_html
        assert 'qrScanner' in dashboard_html

        student_schedule = client.get('/student/schedule')
        assert student_schedule.status_code == 200
        assert 'Ders Programım' in student_schedule.get_data(as_text=True)
        assert 'Ders Programı' in dashboard_html

        blocked_check_in = client.post(
            f'/student/session/{att_session.id}/check-in',
            data={'code': 'ABC123', 'override': '1', 'override_reason': 'Mobil bağlantı'},
            environ_base={'REMOTE_ADDR': '8.8.8.8'},
        )
        assert blocked_check_in.status_code in (302, 303)
        assert db.query(AttendanceRecord).filter_by(session_id=att_session.id, student_id=student.id).first() is None

        pair_student_device(student)

        check_in = client.post(
            f'/student/session/{att_session.id}/check-in',
            data={'code': 'ABC123', 'override': '1', 'override_reason': 'Mobil bağlantı'},
            environ_base={'REMOTE_ADDR': '8.8.8.8'},
        )
        assert check_in.status_code in (302, 303)
        record = db.query(AttendanceRecord).filter_by(session_id=att_session.id, student_id=student.id).first()
        assert record is not None
        assert record.status == 'suspicious'

        client.get('/logout')
        login(client, 'teacher')
        teacher_page = client.get(f'/teacher/session/{att_session.id}')
        assert teacher_page.status_code == 200
        teacher_page_html = teacher_page.get_data(as_text=True)
        assert 'Şüpheli Yoklamalar' in teacher_page_html
        assert 'data:image/png;base64,' in teacher_page_html
        assert f'id="currentCode">{att_session.current_code}</div>' in teacher_page_html

        review = client.post(
            f'/teacher/records/{record.id}/resolve',
            data={'decision': 'approve', 'note': 'Kontrol edildi'},
        )
        assert review.status_code in (302, 303)
        db.refresh(record)
        assert record.status == 'approved'

        export = client.get(f'/teacher/export/course/{course.id}')
        assert export.status_code == 200
        workbook = load_workbook(BytesIO(export.data))
        assert 'Yoklama Kayıtları' in workbook.sheetnames

        client.get('/logout')
        login(client, 'other')
        denied = client.get(f'/teacher/export/course/{course.id}', follow_redirects=False)
        assert denied.status_code in (302, 303)

        client.get('/logout')
        login(client, 'admin')
        assert client.get('/admin/statistics').status_code == 200
        all_export = client.get('/admin/export/all')
        assert all_export.status_code == 200
        all_workbook = load_workbook(BytesIO(all_export.data))
        assert 'Genel Özet' in all_workbook.sheetnames

        verification_session = AttendanceSession(
            course_id=course.id,
            teacher_id=teacher.id,
            current_code='XYZ789',
            code_expires_at=(datetime.utcnow() + timedelta(minutes=5)).isoformat(),
            code_refresh_seconds=10,
            allowed_ip_prefix='10.0.',
        )
        db.add(verification_session)
        db.commit()

        client.get('/logout')
        login(client, 'student')
        manual = client.post('/student/api/manual-verification')
        assert manual.status_code == 200
        start_verification = client.get('/student/api/start-verification')
        assert start_verification.status_code == 200
        verification_check_in = client.post(
            '/student/api/submit-verification',
            json={'session_id': verification_session.id, 'code': 'XYZ789'},
        )
        assert verification_check_in.status_code == 200
        verification_record = db.query(AttendanceRecord).filter_by(
            session_id=verification_session.id,
            student_id=student.id,
        ).first()
        assert verification_record is not None
        assert verification_record.status == 'suspicious'

        expired_session = AttendanceSession(
            course_id=course.id,
            teacher_id=teacher.id,
            current_code='OLD123',
            code_expires_at=(datetime.utcnow() - timedelta(seconds=1)).isoformat(),
            code_refresh_seconds=10,
        )
        db.add(expired_session)
        db.commit()
        expired_record, error = attendance_service.check_in(expired_session.id, student.id, 'OLD123', ip_address='127.0.0.1')
        assert expired_record is None
        assert 'süresi doldu' in error

    print('integration ok')


if __name__ == '__main__':
    run()
