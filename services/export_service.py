from io import BytesIO
import re

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from database import db
from models.attendance_record import AttendanceRecord
from models.attendance_session import AttendanceSession
from models.course import Course, CourseStudent
from models.user import User

PRESENT_STATUSES = ('verified', 'approved', 'manual')

STATUS_TR = {
    'verified': 'Doğrulandı',
    'approved': 'Onaylandı',
    'manual': 'Manuel',
    'suspicious': 'Şüpheli',
    'rejected': 'Reddedildi',
}


def _clean_date(value):
    if not value:
        return '-'
    text = value.isoformat(sep=' ') if hasattr(value, 'isoformat') else str(value)
    if text.startswith('(datetime'):
        return '-'
    return text[:19]


def _safe_filename(value):
    name = re.sub(r'[^A-Za-z0-9._-]+', '_', value or 'yoklama')
    name = name.strip('._-')
    return name or 'yoklama'


def _safe_sheet_title(value, used_titles):
    title = re.sub(r'[\[\]:*?/\\]+', ' ', value or 'Ders').strip()
    title = re.sub(r'\s+', ' ', title) or 'Ders'

    base = title[:31]
    candidate = base
    counter = 2
    while candidate in used_titles:
        suffix = f' {counter}'
        candidate = f'{base[:31 - len(suffix)]}{suffix}'
        counter += 1

    used_titles.add(candidate)
    return candidate


HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
HEADER_FILL = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
HEADER_ALIGN = Alignment(horizontal='center', vertical='center')
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin'),
)


def _style_header(ws, row, col_count):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER


def _auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 40)


def _rate(part, total):
    return round((part / total * 100), 1) if total else 0


def export_course_attendance(course_id):
    course = db.query(Course).filter_by(id=course_id).first()
    if not course:
        return None, 'Ders bulunamadı.'

    wb = Workbook()

    # --- Sheet 1: Oturum Detayi ---
    ws1 = wb.active
    ws1.title = 'Yoklama Kayıtları'
    headers = ['Öğrenci No', 'Ad', 'Oturum Tarihi', 'Durum', 'Kod', 'IP', 'GPS Mesafe (m)', 'Gönderim Zamanı']
    ws1.append(headers)
    _style_header(ws1, 1, len(headers))

    sessions = db.query(AttendanceSession).filter_by(course_id=course_id).order_by(AttendanceSession.started_at).all()
    enrolled = db.query(CourseStudent).filter_by(course_id=course_id).all()
    student_ids = [e.student_id for e in enrolled]
    students_map = {}
    if student_ids:
        students = db.query(User).filter(User.id.in_(student_ids)).all()
        students_map = {s.id: s for s in students}

    for att_session in sessions:
        records = db.query(AttendanceRecord).filter_by(session_id=att_session.id).all()
        submitted_ids = {r.student_id for r in records}

        for record in records:
            student = students_map.get(record.student_id)
            ws1.append([
                student.student_number if student else '-',
                student.username if student else '-',
                _clean_date(att_session.started_at)[:10],
                STATUS_TR.get(record.status, record.status),
                record.submitted_code or '-',
                record.ip_address or '-',
                round(record.gps_distance_m, 1) if record.gps_distance_m is not None else '-',
                _clean_date(record.submitted_at),
            ])

        for sid in student_ids:
            if sid not in submitted_ids:
                student = students_map.get(sid)
                ws1.append([
                    student.student_number if student else '-',
                    student.username if student else '-',
                    _clean_date(att_session.started_at)[:10],
                    'Girilmemis',
                    '-', '-', '-', '-',
                ])

    _auto_width(ws1)

    # --- Sheet 2: Öğrenci Özet ---
    ws2 = wb.create_sheet('Öğrenci Özet')
    headers2 = ['Öğrenci No', 'Ad', 'Bölüm', 'Toplam Oturum', 'Katılım', 'Şüpheli', 'Devamsız', 'Oran (%)']
    ws2.append(headers2)
    _style_header(ws2, 1, len(headers2))

    session_count = len(sessions)
    session_ids = [s.id for s in sessions]

    for sid in student_ids:
        student = students_map.get(sid)
        if not student:
            continue
        records = []
        if session_ids:
            records = db.query(AttendanceRecord).filter(
                AttendanceRecord.student_id == sid,
                AttendanceRecord.session_id.in_(session_ids),
            ).all()
        present = len([r for r in records if r.status in PRESENT_STATUSES])
        suspicious = len([r for r in records if r.status == 'suspicious'])
        submitted = len(records)
        absent = max(session_count - present - suspicious, 0)
        ws2.append([
            student.student_number or '-',
            student.username,
            student.department or '-',
            session_count,
            present,
            suspicious,
            absent,
            _rate(present, session_count),
        ])

    _auto_width(ws2)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f'{_safe_filename(course.code or course.name)}_yoklama.xlsx'
    return buf, filename


def export_all_courses():
    wb = Workbook()
    wb.remove(wb.active)

    courses = db.query(Course).order_by(Course.name).all()
    if not courses:
        ws = wb.create_sheet('Bos')
        ws.append(['Henüz ders bulunmuyor.'])
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf, 'tum_dersler_yoklama.xlsx'

    # --- Genel Ozet sheet ---
    ws_summary = wb.create_sheet('Genel Özet')
    headers = ['Ders', 'Kod', 'Öğretmen', 'Oturum', 'Kayıtlı', 'Beklenen', 'Katılım', 'Şüpheli', 'Devamsız', 'Oran (%)']
    ws_summary.append(headers)
    _style_header(ws_summary, 1, len(headers))

    for course in courses:
        sessions = db.query(AttendanceSession).filter_by(course_id=course.id).all()
        session_count = len(sessions)
        enrolled_count = db.query(CourseStudent).filter_by(course_id=course.id).count()
        expected = session_count * enrolled_count

        records = db.query(AttendanceRecord).filter_by(course_id=course.id).all()
        present = len([r for r in records if r.status in PRESENT_STATUSES])
        suspicious = len([r for r in records if r.status == 'suspicious'])
        submitted = len(records)
        absent = max(expected - present - suspicious, 0)

        teacher = db.query(User).filter_by(id=course.teacher_id).first()

        ws_summary.append([
            course.name,
            course.code or '-',
            teacher.username if teacher else '-',
            session_count,
            enrolled_count,
            expected,
            present,
            suspicious,
            absent,
            _rate(present, expected),
        ])

    _auto_width(ws_summary)

    # --- Her ders icin ayri sheet ---
    used_titles = set(wb.sheetnames)
    for course in courses:
        sheet_name = _safe_sheet_title(course.code or course.name, used_titles)
        ws = wb.create_sheet(sheet_name)
        headers = ['Öğrenci No', 'Ad', 'Bölüm', 'Oturum Sayısı', 'Katılım', 'Şüpheli', 'Devamsız', 'Oran (%)']
        ws.append(headers)
        _style_header(ws, 1, len(headers))

        sessions = db.query(AttendanceSession).filter_by(course_id=course.id).all()
        session_count = len(sessions)
        session_ids = [s.id for s in sessions]

        enrolled = db.query(CourseStudent).filter_by(course_id=course.id).all()
        student_ids = [e.student_id for e in enrolled]
        students = db.query(User).filter(User.id.in_(student_ids)).all() if student_ids else []
        students_map = {s.id: s for s in students}

        for sid in student_ids:
            student = students_map.get(sid)
            if not student:
                continue
            records = []
            if session_ids:
                records = db.query(AttendanceRecord).filter(
                    AttendanceRecord.student_id == sid,
                    AttendanceRecord.session_id.in_(session_ids),
                ).all()
            present = len([r for r in records if r.status in PRESENT_STATUSES])
            suspicious = len([r for r in records if r.status == 'suspicious'])
            absent = max(session_count - present - suspicious, 0)
            ws.append([
                student.student_number or '-',
                student.username,
                student.department or '-',
                session_count,
                present,
                suspicious,
                absent,
                _rate(present, session_count),
            ])

        _auto_width(ws)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf, 'tum_dersler_yoklama.xlsx'
